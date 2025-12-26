from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from config import Config

class ExcelReportGenerator:
    def __init__(self):
        self.reports_dir = "reports"
        os.makedirs(self.reports_dir, exist_ok=True)

        self.header_fill = PatternFill(
            start_color="CCCCCC",
            end_color="CCCCCC",
            fill_type="solid"
        )
        self.header_font = Font(bold=True)
        self.center_alignment = Alignment(horizontal="center", vertical="center")

        self.supplier_colors = {
            "industrialsupply": "FFCCFFCC",  # Светло-зеленый
            "machineparts": "FFF9CB9C",      # Светло-оранжевый
            "factorystock": "FFFFFFCC"       # Светло-желтый
        }

    def generate_report(self, analysis_results, user_info=None):
        """Генерация Excel отчета"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Анализ запчастей"

        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f"Отчет анализа промышленных запчастей\n{datetime.now().strftime('%d.%m.%Y %H:%M')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        if user_info:
            ws['A2'] = f"Пользователь: {user_info.get('username', 'Неизвестно')}"
            ws['B2'] = f"ID: {user_info.get('id', 'Неизвестно')}"

        row_idx = 4

        for result in analysis_results:
            ws.merge_cells(f'A{row_idx}:J{row_idx}')
            part_header = ws[f'A{row_idx}']
            part_header.value = f"Запчасть: {result['part_number']} - {result['name']}"
            part_header.font = Font(bold=True, size=12, color="000000")
            part_header.fill = PatternFill(
                start_color="DDDDDD",
                end_color="DDDDDD",
                fill_type="solid"
            )
            row_idx += 1

            info_rows = [
                ["Бренды", ", ".join(result["brands"])],
                ["Мин. цена", f"{result['min_price']['price']} руб. ({result['min_price']['supplier_name']})"],
                ["Мед. цена", f"{result['median_price']['price']} руб. ({result['median_price']['supplier_name']})"],
                ["Срок доставки", f"{result['min_price']['delivery']} дней (мин.)"]
            ]

            for label, value in info_rows:
                ws[f'A{row_idx}'] = label
                ws[f'B{row_idx}'] = value
                row_idx += 1

            row_idx += 1

            headers = ["Поставщик", "Бренд", "Цена (руб.)", "Срок (дней)", "Примечание"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_alignment

            row_idx += 1

            for price in result["all_prices"]:
                ws.cell(row=row_idx, column=1, value=price["supplier_name"])
                ws.cell(row=row_idx, column=2, value=price["brand"])
                ws.cell(row=row_idx, column=3, value=price["price"])
                ws.cell(row=row_idx, column=4, value=price["delivery"])

                supplier_code = price["supplier"]
                if supplier_code in self.supplier_colors:
                    fill_color = self.supplier_colors[supplier_code]
                    for col in range(1, 6):
                        ws.cell(row=row_idx, column=col).fill = PatternFill(
                            start_color=fill_color,
                            end_color=fill_color,
                            fill_type="solid"
                        )

                if price["price"] == result["min_price"]["price"]:
                    ws.cell(row=row_idx, column=5, value="МИНИМАЛЬНАЯ ЦЕНА")
                    for col in range(1, 6):
                        ws.cell(row=row_idx, column=col).font = Font(
                            bold=True,
                            color="FF8C00"  # Оранжевый
                        )

                elif price["price"] == result["median_price"]["price"]:
                    ws.cell(row=row_idx, column=5, value="МЕДИАННАЯ ЦЕНА")
                    for col in range(1, 6):
                        ws.cell(row=row_idx, column=col).font = Font(
                            bold=True,
                            color="0000FF"  # Синий
                        )

                row_idx += 1

            row_idx += 1
            ws.merge_cells(f'A{row_idx}:E{row_idx}')
            analogs_header = ws[f'A{row_idx}']
            analogs_header.value = "Доступные аналоги:"
            analogs_header.font = Font(bold=True)
            row_idx += 1

            for analog in result["analogs"]:
                ws[f'A{row_idx}'] = analog["part_number"]
                ws[f'B{row_idx}'] = f"~{analog['estimated_price']:.0f} руб."
                ws[f'C{row_idx}'] = analog["availability"]
                row_idx += 1

            row_idx += 2

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        filename = f"parts_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        wb.save(filepath)

        return filepath

report_generator = ExcelReportGenerator()
